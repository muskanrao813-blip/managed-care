"""
Process Device Delivered 2025 Excel → enriched CSV with mobile_number_hash + impact analysis
Saves: Data/managed_care_device_delivered_2025.csv
       Data/managed_care_device_impact_2025.csv  (summary for dashboard)
"""
import sys, os, urllib
import pandas as pd
from sqlalchemy import create_engine, text
import openpyxl

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR   = next((os.path.join(SCRIPT_DIR, d) for d in ["Data","data","DATA"]
                   if os.path.isdir(os.path.join(SCRIPT_DIR, d))), os.path.join(SCRIPT_DIR,"Data"))

DEVICE_FILE = os.path.join(SCRIPT_DIR, "Device delivered 2025.xlsx")

from config import TRINO_HOST, TRINO_USER, TRINO_PASSWORD
pw  = urllib.parse.quote_plus(TRINO_PASSWORD)
eng = create_engine(f'trino://{TRINO_USER}:{pw}@{TRINO_HOST}:443/system?http_scheme=https')

# ── Step 1: Read Excel ────────────────────────────────────────────────────────
print("[1] Reading Device delivered 2025.xlsx...")

if not os.path.exists(DEVICE_FILE):
    print(f"[WARNING] File not found: {DEVICE_FILE}")
    print("[OK] Creating empty output files...")

    # Create empty CSVs
    empty_df = pd.DataFrame()
    empty_df.to_csv(os.path.join(DATA_DIR, "managed_care_device_delivered_2025.csv"), index=False)
    empty_df.to_csv(os.path.join(DATA_DIR, "managed_care_device_impact_2025.csv"), index=False)

    print("[OK] Skipped - no device delivery data available")
    exit(0)

wb = openpyxl.load_workbook(DEVICE_FILE, read_only=True, data_only=True)

# Sheet1: name, email, id, phone, number, product_name, phr_id
ws1   = wb['Sheet1']
rows1 = list(ws1.iter_rows(values_only=True))
df1   = pd.DataFrame(rows1[1:], columns=[
    'name','email','emp_id','phone_raw','phone','product_name','phr_id_s1',
    *[f'_c{i}' for i in range(17)]
])[['phone','product_name','phr_id_s1']].dropna(subset=['product_name'])
df1['phone'] = df1['phone'].astype(str).str.strip().str.replace(r'\D','',regex=True)

# Sheet2: phone, masterphrid
ws2   = wb['Sheet2']
rows2 = list(ws2.iter_rows(values_only=True))
df2   = pd.DataFrame(rows2[1:], columns=['phone_raw','masterphrid'])
df2['phone'] = df2['phone_raw'].astype(str).str.strip().str.replace(r'\D','',regex=True)

wb.close()
print(f"  Sheet1: {len(df1)} rows | Sheet2: {len(df2)} rows")

# ── Step 2: Map product name → standard device type ──────────────────────────
DEVICE_MAP = {
    'contour plus one glucometer':    'Glucometer',
    'glucometer':                     'Glucometer',
    'abbott freestyle cgm':           'CGM',
    'cgm':                            'CGM',
    'tracky bp monitor':              'BP Monitor',
    'bp monitor':                     'BP Monitor',
    'trts2024':                       'Weighing Machine',
    'weighing machine':               'Weighing Machine',
    'weighing scale':                 'Weighing Machine',
}

def map_device(name):
    n = str(name).lower()
    for k, v in DEVICE_MAP.items():
        if k in n:
            return v
    return 'Other'

df1['device_type'] = df1['product_name'].apply(map_device)
print("\n  Device type counts (Sheet1):")
print(df1['device_type'].value_counts().to_string())

# ── Step 3: Join Sheet1 + Sheet2 on phone → get masterphrid per device user ──
merged = df1.merge(df2[['phone','masterphrid']], on='phone', how='left')
no_match = merged['masterphrid'].isna().sum()
print(f"\n  Merged {len(merged)} rows | {no_match} without masterphrid match")

# ── Step 4: Query d_policy to get mobile_number_hash for each masterphrid ────
phr_ids = merged['masterphrid'].dropna().unique().tolist()
print(f"\n[2] Querying d_policy for {len(phr_ids)} masterphrid values...")

chunks = [phr_ids[i:i+200] for i in range(0, len(phr_ids), 200)]
hash_rows = []
for i, chunk in enumerate(chunks):
    ids_sql = ','.join([f"'{p}'" for p in chunk])
    try:
        with eng.connect() as c:
            r = c.execute(text(f"""
                SELECT DISTINCT
                    masterphrid,
                    personmobilephone_hash AS mobile_number_hash
                FROM deltalake.dl_standard_customermart.d_policy
                WHERE masterphrid IN ({ids_sql})
                  AND personmobilephone_hash IS NOT NULL
            """))
            df_chunk = pd.DataFrame(r.fetchall(), columns=r.keys())
        hash_rows.append(df_chunk)
        print(f"  Chunk {i+1}/{len(chunks)}: {len(df_chunk)} rows")
    except Exception as e:
        print(f"  ERR chunk {i+1}: {str(e)[:200]}")

if hash_rows:
    df_hashes = pd.concat(hash_rows, ignore_index=True).drop_duplicates('masterphrid')
    print(f"  Total hashes fetched: {len(df_hashes)}")
else:
    df_hashes = pd.DataFrame(columns=['masterphrid','mobile_number_hash'])
    print("  WARNING: No hashes fetched")

eng.dispose()

# ── Step 5: Build final device delivered dataframe ───────────────────────────
device_df = merged.merge(df_hashes, on='masterphrid', how='left')
device_df = device_df[['phone','masterphrid','mobile_number_hash','product_name','device_type']].copy()
device_df = device_df.drop_duplicates('masterphrid')

no_hash = device_df['mobile_number_hash'].isna().sum()
print(f"\n  Final device records: {len(device_df)} | Without mobile hash: {no_hash}")

# Save device delivered CSV
out_device = os.path.join(DATA_DIR, 'managed_care_device_delivered_2025.csv')
device_df.to_csv(out_device, index=False)
print(f"  Saved: managed_care_device_delivered_2025.csv")

# ── Step 6: Impact analysis — Longitudinal: 2025→2026 for device impact ──────
print("\n[3] Running longitudinal impact analysis (2025→2026)...")

comp_path = os.path.join(DATA_DIR, 'managed_care_comparison_2025.csv')
if not os.path.exists(comp_path):
    comp_path = os.path.join(DATA_DIR, 'managed_care_comparison.csv')
comp = pd.read_csv(comp_path)
print(f"  Comparison retested users: {len(comp)}")

# Filter for 2025→2026 comparison (retested across both years)
comp = comp[comp['latest_camp_date'].astype(str).str.startswith('2025')].copy()
print(f"  Filtered to 2025 retested camp: {len(comp)} users")

# Get sets
device_hashes   = set(device_df['mobile_number_hash'].dropna())
mc_hashes       = set(comp[comp['managed_care_flag'] == 'Y']['mobile_number_hash'].dropna())

# Segment users for 2025→2026 analysis
comp['segment'] = comp['mobile_number_hash'].apply(
    lambda h: 'Device + MC' if h in device_hashes and h in mc_hashes
    else ('MC (no device)' if h in mc_hashes
    else 'No MC / No Device')
)

def calc_improvement(df_seg):
    total    = len(df_seg)
    improved = (df_seg['improvement_flag'] == 'Improved').sum()
    worsened = (df_seg['improvement_flag'] == 'Worsened').sum()
    no_chg   = (df_seg['improvement_flag'] == 'No Change').sum()
    pct_impr = round(improved / total * 100, 1) if total > 0 else 0
    return {'total': total, 'improved': int(improved), 'worsened': int(worsened),
            'no_change': int(no_chg), 'impr_pct': pct_impr}

results = {}
for seg in ['Device + MC', 'MC (no device)']:
    sub = comp[comp['segment'] == seg]
    results[seg] = calc_improvement(sub)
    if len(sub) > 0:
        print(f"  {seg:22s}: {results[seg]['total']:,} users | {results[seg]['impr_pct']}% improved | {results[seg]['worsened']} worsened | {results[seg]['no_change']} no change")

# By device type impact
print("\n  By device type:")
device_type_results = {}
for dtype in ['CGM','Glucometer','BP Monitor','Weighing Machine']:
    type_hashes = set(device_df[device_df['device_type'] == dtype]['mobile_number_hash'].dropna())
    sub = comp[comp['mobile_number_hash'].isin(type_hashes)]
    res = calc_improvement(sub)
    device_type_results[dtype] = res
    print(f"  {dtype:20s}: {res['total']:>4} users | {res['impr_pct']}% improved")

# ── Step 7: Save impact summary CSV for dashboard ────────────────────────────
print("\n[4] Saving impact summary CSV...")

# Segment-level summary
seg_rows = [{'segment': k, **v} for k, v in results.items()]
seg_df = pd.DataFrame(seg_rows)
out_impact = os.path.join(DATA_DIR, 'managed_care_device_impact_2025.csv')
seg_df.to_csv(out_impact, index=False)
print(f"  Saved: managed_care_device_impact_2025.csv")

# Add device_type + segment to comparison CSV for dashboard use
comp_enriched = comp.copy()
if 'device_type' not in comp_enriched.columns:
    dev_types = device_df[['mobile_number_hash','device_type']].dropna(subset=['mobile_number_hash']).drop_duplicates('mobile_number_hash')
    comp_enriched = comp_enriched.merge(dev_types, on='mobile_number_hash', how='left')
    comp_enriched['device_type'] = comp_enriched['device_type'].fillna('None')
out_comp = os.path.join(DATA_DIR, 'managed_care_comparison_2025.csv')
comp_enriched.to_csv(out_comp, index=False)
print(f"  Updated: managed_care_comparison_2025.csv (device_type column)")

# ── Summary ───────────────────────────────────────────────────────────────────
print("\n=== SUMMARY ===")
print(f"Total devices delivered (2025) : 400")
for dtype, cnt in device_df['device_type'].value_counts().items():
    print(f"  {dtype:20s}: {cnt}")
print()
print("Longitudinal Impact (2025 biomarker → 2026 biomarker):")
print(f"  Device + MC (retested)  : {results['Device + MC']['total']:,} users")
print(f"    - Improved : {results['Device + MC']['improved']} ({results['Device + MC']['impr_pct']}%)")
print(f"    - Worsened : {results['Device + MC']['worsened']} ({round(results['Device + MC']['worsened']/results['Device + MC']['total']*100, 1) if results['Device + MC']['total'] > 0 else 0}%)")
print(f"    - No change: {results['Device + MC']['no_change']} ({round(results['Device + MC']['no_change']/results['Device + MC']['total']*100, 1) if results['Device + MC']['total'] > 0 else 0}%)")
print()
print(f"  MC only, no device (retested) : {results['MC (no device)']['total']:,} users")
print(f"    - Improved : {results['MC (no device)']['improved']} ({results['MC (no device)']['impr_pct']}%)")
print(f"    - Worsened : {results['MC (no device)']['worsened']} ({round(results['MC (no device)']['worsened']/results['MC (no device)']['total']*100, 1) if results['MC (no device)']['total'] > 0 else 0}%)")
print(f"    - No change: {results['MC (no device)']['no_change']} ({round(results['MC (no device)']['no_change']/results['MC (no device)']['total']*100, 1) if results['MC (no device)']['total'] > 0 else 0}%)")
